# imports
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import io
import tempfile
from openpyxl.worksheet.table import Table, TableStyleInfo

# logo path
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "assets", "nhs_logo.jpeg")

# Excel colours (aRGB format for openpyxl)
NHS_BLUE = "FF005EB8"
WHITE = "FFFFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=NHS_BLUE)
HEADER_FONT = Font(bold=True, color=WHITE)


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def add_dataframe_sheet(wb, name, df):
    ws = wb.create_sheet(name)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    style_header(ws)

    # Format numbers to 2 decimals & add comma separators
    for col in ws.iter_cols(min_row=2):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    # Auto width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2
        for cell in col[1:]:
            if isinstance(cell.value, float):
                cell.value = round(cell.value, 2)

    # Add Excel table
    tab = Table(displayName=f"{name.replace(' ', '_')}_table",ref=f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    return ws


def add_logo(ws, cell="A1"):
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width = 120
        img.height = 60
        ws.add_image(img, cell)
    else:
        ws[cell] = "NHS Logo missing"


def add_plotly_figure(ws, fig, cell="H2"):
    import tempfile
    import os

    tmp_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp_file.close()

    fig.write_image(tmp_file.name)
    img = XLImage(tmp_file.name)
    img.width = 500
    img.height = 300
    ws.add_image(img, cell)

    # Cleanup after saving workbook
    return tmp_file.name





def to_excel_report(combined_df, hospital_summary, comparison_df, fig1, fig2, assumptions):
    import io
    import os
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    temp_files = []

    # GP sheet
    ws1 = add_dataframe_sheet(wb, "GP Analysis", combined_df)
    temp_files.append(add_plotly_figure(ws1, fig1))

    # Hospital sheet
    ws2 = add_dataframe_sheet(wb, "Hospital Summary", hospital_summary)
    temp_files.append(add_plotly_figure(ws2, fig2))

    # Assumptions sheet
    ws3 = wb.create_sheet("Model Assumptions")
    ws3.append(["Parameter", "Value"])
    style_header(ws3)
    for k, v in assumptions.items():
        ws3.append([k, v])

  
    # Scenario comparison sheet
    ws4 = add_dataframe_sheet(wb, "Scenario Comparison", comparison_df)


    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Cleanup temp image files
    for f in temp_files:
        os.remove(f)

    return output.getvalue()
